import shutil
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json

# Copy all PDF page images to a folder next to the Excel for user reference
img_src = Path('/Users/shiqiuxun/Documents/日常/workbench/pdf_pages_rot')
img_dst = Path('/Users/shiqiuxun/Desktop/PDF扫描页证据')
if img_dst.exists():
    shutil.rmtree(img_dst)
img_dst.mkdir()

for p in sorted(img_src.glob('*.png')):
    shutil.copy(p, img_dst / p.name)
print(f'Copied {len(list(img_dst.glob("*.png")))} page images to {img_dst}')

# Now update the Excel to add hyperlinks
dst = '/Users/shiqiuxun/Desktop/资产盘点明细表-2026-07-25更新版.xlsx'
wb = load_workbook(dst)

link_font = Font(color='0000FF', underline='single')

for sheet_name in wb.sheetnames:
    if sheet_name == '扫描情况总览':
        continue
    ws = wb[sheet_name]
    # Find PDF页码 column
    pdf_col = None
    header_row = 1
    for hr in range(1, 6):
        for c in range(1, ws.max_column+1):
            v = str(ws.cell(row=hr, column=c).value or '')
            if v == 'PDF页码':
                pdf_col = c
                header_row = hr
                break
        if pdf_col:
            break

    if pdf_col is None:
        continue

    for row in range(header_row+1, ws.max_row+1):
        cell = ws.cell(row=row, column=pdf_col)
        val = str(cell.value or '').strip()
        if val and val.startswith('p'):
            # Set hyperlink to the corresponding image file
            first_page = val.split(',')[0].strip()  # e.g., "p42"
            try:
                img_name = f'PDF扫描页证据/{first_page}.png'
                cell.hyperlink = img_name
                cell.font = link_font
            except Exception as e:
                print(f'Error on {sheet_name} row {row}: {e}')

# Also add hyperlinks to summary sheet
sum_ws = wb['扫描情况总览']
# Save and verify
wb.save(dst)
print(f'Excel saved with hyperlinks to {dst}')
print(f'PDF page images at: {img_dst}')
