from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import os
from PIL import Image as PILImage

dst = '/Users/shiqiuxun/Desktop/资产盘点明细表-2026-07-25更新版.xlsx'
wb = load_workbook(dst)
ws = wb['扫描情况总览']

# Create a comprehensive image showing what got matched vs missing
# Just embed the overview thumbnail
overview = '/Users/shiqiuxun/Documents/日常/workbench/all_pages_overview.png'
if os.path.exists(overview):
    # Resize the overview first to embed at reasonable size
    img = PILImage.open(overview)
    img.thumbnail((900, 1700))
    tmp = '/Users/shiqiuxun/Documents/日常/workbench/overview_small.png'
    img.save(tmp)
    xl_img = XLImage(tmp)
    # Place at F2
    ws.add_image(xl_img, 'F2')

# Also add a row showing the folder location
max_row = ws.max_row
ws.cell(row=max_row+2, column=1, value='📂 文件位置')
ws.cell(row=max_row+2, column=1).font = Font(bold=True, size=12)

ws.cell(row=max_row+3, column=1, value='更新版 Excel：')
ws.cell(row=max_row+3, column=2, value=dst)

ws.cell(row=max_row+4, column=1, value='PDF 扫描证据图（63 页）：')
ws.cell(row=max_row+4, column=2, value='/Users/shiqiuxun/Desktop/PDF扫描页证据/')

ws.cell(row=max_row+5, column=1, value='原始 PDF：')
ws.cell(row=max_row+5, column=2, value='/Users/shiqiuxun/Desktop/扫描图片_2026-07-25.pdf')

ws.cell(row=max_row+6, column=1, value='原始 Excel（参考）：')
ws.cell(row=max_row+6, column=2, value='/Users/shiqiuxun/Desktop/附件一：资产盘点明细表-2025-09-20.xlsx')

# Use next section
ws.cell(row=max_row+8, column=1, value='📖 使用说明')
ws.cell(row=max_row+8, column=1).font = Font(bold=True, size=12)

instructions = [
    ('①', '打开工作簿，第一张表"扫描情况总览"显示了匹配统计与 PDF 全 63 页缩略图。'),
    ('②', '每个工作表末尾追加了 5 列：PDF页码｜扫描状态｜扫描存放地点｜扫描盘点数｜差异说明'),
    ('③', '黄色高亮的行代表对应资产在 PDF 中找到了。请点击该行的 PDF 页码（蓝色下划线），会自动跳到 PDF扫描页证据文件夹打开对应图片。'),
    ('④', '对红线勾选（✓）的资产，确认即可；对有手写改动（损坏/不见/搬移/数量变化）的，请在"差异说明"列填写调整依据，必要时更新"扫描存放地点/扫描盘点数"列。'),
    ('⑤', 'OCR 失败页：p15, p46, p51-p57（多为手写汇总）。请人工打开 PDF 这些页查看。'),
    ('⑥', '⚠️ 本工具未自动修改原数据。所有改动需人工根据 PDF 红✓和手写内容填写。'),
]
r = max_row+9
for mark, text in instructions:
    ws.cell(row=r, column=1, value=mark).font = Font(bold=True)
    ws.cell(row=r, column=2, value=text)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 32
    r += 1

ws.column_dimensions['B'].width = 65

wb.save(dst)
print(f'Updated summary sheet with overview image and instructions')
print(f'Saved: {dst}')
