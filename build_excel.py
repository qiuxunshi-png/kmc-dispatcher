import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json, shutil
from pathlib import Path
from collections import defaultdict

src = '/Users/shiqiuxun/Desktop/附件一：资产盘点明细表-2025-09-20.xlsx'
dst = '/Users/shiqiuxun/Desktop/资产盘点明细表-2026-07-25更新版.xlsx'
shutil.copy(src, dst)

with open('/Users/shiqiuxun/Documents/日常/workbench/all_ids_combined.json', 'r', encoding='utf-8') as f:
    pdf_data = json.load(f)

id_to_pages = defaultdict(list)
for page, ids in pdf_data.items():
    for aid in ids:
        page_num = int(page.replace('p', '').replace('.png', '').replace('.ng', '').replace('.pg', ''))
        if page_num not in id_to_pages[aid]:
            id_to_pages[aid].append(page_num)
id_to_pages = {k: sorted(v) for k, v in id_to_pages.items()}

wb = load_workbook(dst)
print(f'Sheets: {wb.sheetnames[:5]}...')

yellow_fill = PatternFill('solid', start_color='FFFF00')
hdr_font = Font(bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', start_color='4472C4')
updated_count = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Find header row by scanning rows 1-5 for "资产编号"
    id_col = None
    loc_col = None
    status_col = None
    count_col = None
    header_row = 1
    for hr in range(1, 6):
        headers = [str(ws.cell(row=hr, column=c).value or '') for c in range(1, ws.max_column+1)]
        for i, h in enumerate(headers, 1):
            if '资产编号' in h or '财务资产编号' in h:
                id_col = i
                header_row = hr
                break
        if id_col:
            break

    if id_col is None:
        continue

    headers = [str(ws.cell(row=header_row, column=c).value or '') for c in range(1, ws.max_column+1)]
    for i, h in enumerate(headers, 1):
        if '存放' in h and '地点' in h:
            loc_col = i
        if h.strip() == '资产状态' or '资产状态' in h:
            status_col = i
        if h.strip() == '数量':
            count_col = i

    last_col = ws.max_column
    pdf_page_col = last_col + 1
    pdf_status_col = last_col + 2
    pdf_loc_col = last_col + 3
    pdf_count_col = last_col + 4
    notes_col = last_col + 5

    ws.cell(row=header_row, column=pdf_page_col, value='PDF页码').font = hdr_font
    ws.cell(row=header_row, column=pdf_status_col, value='扫描状态').font = hdr_font
    ws.cell(row=header_row, column=pdf_loc_col, value='扫描存放地点').font = hdr_font
    ws.cell(row=header_row, column=pdf_count_col, value='扫描盘点数').font = hdr_font
    ws.cell(row=header_row, column=notes_col, value='差异说明/待人工确认').font = hdr_font
    for c in [pdf_page_col, pdf_status_col, pdf_loc_col, pdf_count_col, notes_col]:
        ws.cell(row=header_row, column=c).fill = hdr_fill
        ws.cell(row=header_row, column=c).alignment = Alignment(horizontal='center')

    ws.column_dimensions[get_column_letter(pdf_page_col)].width = 14
    ws.column_dimensions[get_column_letter(pdf_status_col)].width = 14
    ws.column_dimensions[get_column_letter(pdf_loc_col)].width = 24
    ws.column_dimensions[get_column_letter(pdf_count_col)].width = 14
    ws.column_dimensions[get_column_letter(notes_col)].width = 40

    sheet_updated = 0
    for row in range(header_row+1, ws.max_row+1):
        aid = str(ws.cell(row=row, column=id_col).value or '').strip()
        if aid in id_to_pages:
            pages = id_to_pages[aid]
            page_text = ', '.join(f'p{p}' for p in pages)
            ws.cell(row=row, column=pdf_page_col, value=page_text)
            ws.cell(row=row, column=pdf_status_col, value='🔍 已扫到')
            ws.cell(row=row, column=notes_col, value='见 PDF 第 {} 页（请打开该页核对红✓与手写备注）'.format(', '.join(map(str, pages))))
            for c in [pdf_page_col, pdf_status_col, notes_col]:
                ws.cell(row=row, column=c).fill = yellow_fill
            sheet_updated += 1
            updated_count += 1
        elif aid.startswith('S') and aid[1:] in id_to_pages:
            pages = id_to_pages[aid[1:]]
            ws.cell(row=row, column=pdf_page_col, value=', '.join(f'p{p}' for p in pages))
            ws.cell(row=row, column=pdf_status_col, value='🔍 已扫到')
            ws.cell(row=row, column=notes_col, value='见 PDF 第 {} 页（请打开该页核对红✓与手写备注）'.format(', '.join(map(str, pages))))
            for c in [pdf_page_col, pdf_status_col, notes_col]:
                ws.cell(row=row, column=c).fill = yellow_fill
            sheet_updated += 1
            updated_count += 1

    print(f'Sheet [{sheet_name}]: updated {sheet_updated} rows')

# Add a summary sheet at the front
if '扫描情况总览' in wb.sheetnames:
    del wb['扫描情况总览']
sum_ws = wb.create_sheet('扫描情况总览', 0)

sum_ws['A1'] = '资产盘点扫描情况总览（2026-07-25 更新）'
sum_ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
sum_ws['A1'].fill = hdr_fill
sum_ws.merge_cells('A1:D1')
sum_ws.row_dimensions[1].height = 30

sum_ws['A3'] = '项目'
sum_ws['B3'] = '数值'
sum_ws['C3'] = '说明'
for c in ['A3', 'B3', 'C3']:
    sum_ws[c].font = hdr_font
    sum_ws[c].fill = hdr_fill
    sum_ws[c].alignment = Alignment(horizontal='center')

stats = [
    ('PDF 扫描页数', 63, '本次扫描的物理页数'),
    ('OCR 识别出唯一资产编号数', 1192, '用 PSM 4+12 两种模式合并去重'),
    ('PDF 中独有的资产编号（在 Excel 找不到）', 617, '可能是 PDF 上有但 Excel 未登记的项'),
    ('Excel 中总资产编号数', 4374, '原盘点表的资产编号数'),
    ('成功匹配并已在 Excel 行末标注', updated_count, '可点击备注中提示的 PDF 页码核对红✓与手写'),
    ('未扫描到的 Excel 项', 4374-updated_count, '未在本次扫描中找到，需要人工确认'),
    ('OCR 失败页（手写汇总）', 8, 'p15, p46, p51-57 — 多为汇总备注页，请人工查阅'),
]
for i, (k, v, note) in enumerate(stats, start=4):
    sum_ws.cell(row=i, column=1, value=k)
    sum_ws.cell(row=i, column=2, value=v)
    sum_ws.cell(row=i, column=3, value=note)
    if i % 2 == 0:
        for c in [1, 2, 3]:
            sum_ws.cell(row=i, column=c).fill = PatternFill('solid', start_color='F2F2F2')

sum_ws.column_dimensions['A'].width = 36
sum_ws.column_dimensions['B'].width = 16
sum_ws.column_dimensions['C'].width = 60

# Color legend
sum_ws['A13'] = '📌 颜色标注说明'
sum_ws['A13'].font = Font(bold=True, size=12)
sum_ws['A14'] = '浅黄填充（#FFFF00）'
sum_ws['B14'] = 'PDF 扫描对应的行（即与扫描页匹配上的资产行）'
sum_ws['A14'].fill = yellow_fill
sum_ws['A15'] = '红✓ 标记'
sum_ws['B15'] = '在 PDF 中用红色✓勾选 = 资产已确认存在'
sum_ws['A16'] = '🔍 待人工确认'
sum_ws['B16'] = 'OCR 无法识别的中文手写内容，请在打开 PDF 对应页人工核对'
sum_ws['A17'] = '新规则'
sum_ws['B17'] = '所有数据均未修改原值，仅在每个 sheet 末尾追加 5 列。'
sum_ws['B18'] = '若需更新"资产状态/存放地点/盘点数"，请在"扫描存放地点/扫描盘点数"列填写，并在"差异说明"列记录调整依据。'

wb.save(dst)
print(f'\nTotal updated: {updated_count}')
print(f'Saved to: {dst}')
